import math
import re
from collections import Counter
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile, is_zipfile
from datetime import datetime, timezone
import xlrd
from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from models import (
    BranchPriceOverride,
    PriceImportBatch,
    PriceImportRow,
    Product,
)
from services.branch_service import get_branch_by_id

MAXIMUM_FILE_SIZE = 10 * 1024 * 1024
MAXIMUM_UNCOMPRESSED_SIZE = 100 * 1024 * 1024
MAXIMUM_ARCHIVE_FILES = 5000
MAXIMUM_IMPORT_ROWS = 50_000
MAXIMUM_SCANNED_ROWS = 100_000
HEADER_SEARCH_ROWS = 20

TWO_DECIMAL_PLACES = Decimal("0.01")


HEADER_ALIASES = {
    "barcode": {
        "barcode",
        "barcodeid",
        "barcodeno",
        "itemcode",
        "productcode",
        "code",
        "sku",
        "skucode",
    },
    "item_name": {
        "itemname",
        "productname",
        "name",
        "description",
        "discription",
        "itemdescription",
        "itemdiscription",
        "productdescription",
    },
    "price": {
        "price",
        "saleprice",
        "salesprice",
        "sellingprice",
        "retailprice",
        "unitprice",
        "rate",
    },
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""

    return re.sub(
        r"[^a-z0-9]",
        "",
        str(value).strip().lower(),
    )


def detect_header_columns(
    values: list[Any],
) -> dict[str, int] | None:
    detected_columns: dict[str, int] = {}

    for column_index, value in enumerate(values):
        normalized_value = normalize_header(value)

        for field_name, aliases in HEADER_ALIASES.items():
            if (
                field_name not in detected_columns
                and normalized_value in aliases
            ):
                detected_columns[field_name] = column_index

    required_fields = {
        "barcode",
        "item_name",
        "price",
    }

    if required_fields.issubset(detected_columns):
        return detected_columns

    return None


def normalize_barcode(
    value: Any,
    number_format: str | None = None,
) -> str | None:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, int):
        barcode = str(value)

    elif isinstance(value, float):
        if not math.isfinite(value):
            return None

        if value.is_integer():
            barcode = str(int(value))
        else:
            barcode = format(value, ".15g")

    else:
        barcode = str(value).strip()

        if re.fullmatch(r"\d+\.0", barcode):
            barcode = barcode[:-2]

    if not barcode:
        return None

    if number_format:
        first_format = number_format.split(";")[0].strip()

        if (
            re.fullmatch(r"0+", first_format)
            and barcode.isdigit()
        ):
            barcode = barcode.zfill(len(first_format))

    if len(barcode) > 100:
        return None

    return barcode


def normalize_item_name(value: Any) -> str | None:
    if value is None:
        return None

    item_name = str(value).strip()

    if not item_name or len(item_name) > 255:
        return None

    return item_name


def normalize_price(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None

    price_text = str(value).strip().upper()

    replacements = [
        "PKR",
        "RS.",
        "RS",
        "₨",
        ",",
        "/-",
        " ",
    ]

    for replacement in replacements:
        price_text = price_text.replace(replacement, "")

    try:
        price = Decimal(price_text)

    except (InvalidOperation, ValueError):
        return None

    if not price.is_finite() or price < 0:
        return None

    return price.quantize(
        TWO_DECIMAL_PLACES,
        rounding=ROUND_HALF_UP,
    )


def validate_xlsx_archive(file_content: bytes) -> None:
    file_stream = BytesIO(file_content)

    if not is_zipfile(file_stream):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded XLSX file is invalid.",
        )

    try:
        file_stream.seek(0)

        with ZipFile(file_stream) as archive:
            archive_files = archive.infolist()

            if len(archive_files) > MAXIMUM_ARCHIVE_FILES:
                raise HTTPException(
                    status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                    detail="The Excel archive contains too many files.",
                )

            uncompressed_size = sum(
                archive_file.file_size
                for archive_file in archive_files
            )

            if uncompressed_size > MAXIMUM_UNCOMPRESSED_SIZE:
                raise HTTPException(
                    status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                    detail="The uncompressed Excel file is too large.",
                )

    except BadZipFile as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded XLSX file is invalid.",
        ) from error


def extract_xlsx_rows(
    file_content: bytes,
) -> list[dict[str, Any]]:
    validate_xlsx_archive(file_content)

    try:
        workbook = load_workbook(
            BytesIO(file_content),
            read_only=True,
            data_only=True,
        )

    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The XLSX workbook could not be opened.",
        ) from error

    try:
        selected_worksheet = None
        header_row_number = None
        header_columns = None

        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=min(
                    HEADER_SEARCH_ROWS,
                    worksheet.max_row,
                ),
            ):
                values = [cell.value for cell in row]
                detected = detect_header_columns(values)

                if detected:
                    selected_worksheet = worksheet
                    header_row_number = row[0].row
                    header_columns = detected
                    break

            if selected_worksheet is not None:
                break

        if (
            selected_worksheet is None
            or header_row_number is None
            or header_columns is None
        ):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "Required columns were not found. "
                    "The file must contain barcode/item code, "
                    "item name/description and sale price."
                ),
            )

        if (
            selected_worksheet.max_row - header_row_number
            > MAXIMUM_SCANNED_ROWS
        ):
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail="The worksheet contains too many rows.",
            )

        extracted_rows: list[dict[str, Any]] = []

        for row in selected_worksheet.iter_rows(
            min_row=header_row_number + 1
        ):
            barcode_cell = row[header_columns["barcode"]]
            name_cell = row[header_columns["item_name"]]
            price_cell = row[header_columns["price"]]

            values = [
                barcode_cell.value,
                name_cell.value,
                price_cell.value,
            ]

            if all(
                value is None or str(value).strip() == ""
                for value in values
            ):
                continue

            extracted_rows.append(
                {
                    "excel_row_number": barcode_cell.row,
                    "barcode": normalize_barcode(
                        barcode_cell.value,
                        getattr(
                            barcode_cell,
                            "number_format",
                            None,
                        ),
                    ),
                    "item_name": normalize_item_name(
                        name_cell.value
                    ),
                    "uploaded_price": normalize_price(
                        price_cell.value
                    ),
                }
            )

            if len(extracted_rows) > MAXIMUM_IMPORT_ROWS:
                raise HTTPException(
                    status_code=(
                        status.HTTP_413_REQUEST_ENTITY_TOO_LARGE
                    ),
                    detail=(
                        "The Excel file contains more than "
                        "50,000 product rows."
                    ),
                )

        return extracted_rows

    finally:
        workbook.close()


def get_xls_number_format(
    workbook: Any,
    cell: Any,
) -> str | None:
    try:
        format_key = workbook.xf_list[
            cell.xf_index
        ].format_key

        return workbook.format_map[format_key].format_str

    except (AttributeError, IndexError, KeyError):
        return None


def extract_xls_rows(
    file_content: bytes,
) -> list[dict[str, Any]]:
    expected_signature = bytes.fromhex(
        "D0CF11E0A1B11AE1"
    )

    if not file_content.startswith(expected_signature):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded XLS file is invalid.",
        )

    try:
        workbook = xlrd.open_workbook(
            file_contents=file_content,
            formatting_info=True,
            on_demand=True,
        )

    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The XLS workbook could not be opened.",
        ) from error

    try:
        selected_sheet = None
        header_row_index = None
        header_columns = None

        for sheet in workbook.sheets():
            rows_to_check = min(
                HEADER_SEARCH_ROWS,
                sheet.nrows,
            )

            for row_index in range(rows_to_check):
                values = sheet.row_values(row_index)
                detected = detect_header_columns(values)

                if detected:
                    selected_sheet = sheet
                    header_row_index = row_index
                    header_columns = detected
                    break

            if selected_sheet is not None:
                break

        if (
            selected_sheet is None
            or header_row_index is None
            or header_columns is None
        ):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "Required columns were not found. "
                    "The file must contain barcode/item code, "
                    "item name/description and sale price."
                ),
            )

        if (
            selected_sheet.nrows - header_row_index
            > MAXIMUM_SCANNED_ROWS
        ):
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail="The worksheet contains too many rows.",
            )

        extracted_rows: list[dict[str, Any]] = []

        for row_index in range(
            header_row_index + 1,
            selected_sheet.nrows,
        ):
            barcode_cell = selected_sheet.cell(
                row_index,
                header_columns["barcode"],
            )

            name_cell = selected_sheet.cell(
                row_index,
                header_columns["item_name"],
            )

            price_cell = selected_sheet.cell(
                row_index,
                header_columns["price"],
            )

            values = [
                barcode_cell.value,
                name_cell.value,
                price_cell.value,
            ]

            if all(
                value is None or str(value).strip() == ""
                for value in values
            ):
                continue

            extracted_rows.append(
                {
                    "excel_row_number": row_index + 1,
                    "barcode": normalize_barcode(
                        barcode_cell.value,
                        get_xls_number_format(
                            workbook,
                            barcode_cell,
                        ),
                    ),
                    "item_name": normalize_item_name(
                        name_cell.value
                    ),
                    "uploaded_price": normalize_price(
                        price_cell.value
                    ),
                }
            )

            if len(extracted_rows) > MAXIMUM_IMPORT_ROWS:
                raise HTTPException(
                    status_code=(
                        status.HTTP_413_REQUEST_ENTITY_TOO_LARGE
                    ),
                    detail=(
                        "The Excel file contains more than "
                        "50,000 product rows."
                    ),
                )

        return extracted_rows

    finally:
        workbook.release_resources()


def extract_excel_rows(
    file_content: bytes,
    file_extension: str,
) -> list[dict[str, Any]]:
    if file_extension == ".xlsx":
        return extract_xlsx_rows(file_content)

    if file_extension == ".xls":
        return extract_xls_rows(file_content)

    raise HTTPException(
        status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
        detail="Only .xlsx and .xls Excel files are supported.",
    )


def get_products_by_barcodes(
    db: Session,
    barcodes: set[str],
) -> dict[str, Product]:
    products_by_barcode: dict[str, Product] = {}
    barcode_list = list(barcodes)
    chunk_size = 1000

    for start in range(0, len(barcode_list), chunk_size):
        chunk = barcode_list[start : start + chunk_size]

        products = db.scalars(
            select(Product).where(
                Product.barcode.in_(chunk)
            )
        ).all()

        for product in products:
            products_by_barcode[product.barcode] = product

    return products_by_barcode


def get_price_import_preview(
    db: Session,
    batch_id: int,
) -> PriceImportBatch:
    batch = db.scalar(
        select(PriceImportBatch)
        .options(
            selectinload(PriceImportBatch.rows)
        )
        .where(PriceImportBatch.id == batch_id)
    )

    if not batch:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Price import preview not found.",
        )

    return batch


async def create_master_price_preview(
    db: Session,
    excel_file: UploadFile,
) -> PriceImportBatch:
    try:
        original_filename = Path(
            excel_file.filename or "price-import.xlsx"
        ).name

        file_extension = Path(
            original_filename
        ).suffix.lower()

        if file_extension not in {".xlsx", ".xls"}:
            raise HTTPException(
                status_code=(
                    status.HTTP_415_UNSUPPORTED_MEDIA_TYPE
                ),
                detail=(
                    "Only .xlsx and .xls Excel files are supported."
                ),
            )

        file_content = await excel_file.read(
            MAXIMUM_FILE_SIZE + 1
        )

        if not file_content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The uploaded Excel file is empty.",
            )

        if len(file_content) > MAXIMUM_FILE_SIZE:
            raise HTTPException(
                status_code=(
                    status.HTTP_413_REQUEST_ENTITY_TOO_LARGE
                ),
                detail="Excel file size cannot exceed 10 MB.",
            )

        extracted_rows = extract_excel_rows(
            file_content=file_content,
            file_extension=file_extension,
        )

        if not extracted_rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No product rows were found in the Excel file.",
            )

        barcode_counts = Counter(
            row["barcode"]
            for row in extracted_rows
            if row["barcode"]
        )

        valid_barcodes = {
            row["barcode"]
            for row in extracted_rows
            if (
                row["barcode"]
                and row["item_name"]
                and row["uploaded_price"] is not None
                and barcode_counts[row["barcode"]] == 1
            )
        }

        products_by_barcode = get_products_by_barcodes(
            db=db,
            barcodes=valid_barcodes,
        )

        batch = PriceImportBatch(
            import_scope="master",
            branch_id=None,
            original_filename=original_filename[:255],
            status="preview",
        )

        db.add(batch)
        db.flush()

        changed_rows = 0
        unchanged_rows = 0
        invalid_rows = 0

        for extracted_row in extracted_rows:
            barcode = extracted_row["barcode"]
            item_name = extracted_row["item_name"]
            uploaded_price = extracted_row["uploaded_price"]

            row_status = "invalid"
            error_message = None
            apply_selected = False
            product = None
            current_price = None

            validation_errors = []

            if barcode is None:
                validation_errors.append("Invalid or missing barcode.")

            if item_name is None:
                validation_errors.append("Invalid or missing item name.")

            if uploaded_price is None:
                validation_errors.append("Invalid or missing price.")

            if (
                barcode
                and barcode_counts[barcode] > 1
            ):
                validation_errors.append(
                    "Duplicate barcode exists in the uploaded file."
                )

            if validation_errors:
                error_message = " ".join(validation_errors)
                invalid_rows += 1

            else:
                product = products_by_barcode.get(barcode)

                if product is None:
                    row_status = "product_not_found"
                    error_message = (
                        "No existing product was found for this barcode."
                    )
                    invalid_rows += 1

                else:
                    current_price = Decimal(
                        product.master_price
                    ).quantize(TWO_DECIMAL_PLACES)

                    if current_price == uploaded_price:
                        row_status = "unchanged"
                        unchanged_rows += 1

                    else:
                        row_status = "changed"
                        apply_selected = True
                        changed_rows += 1

            import_row = PriceImportRow(
                batch_id=batch.id,
                excel_row_number=(
                    extracted_row["excel_row_number"]
                ),
                product_id=(
                    product.id if product else None
                ),
                barcode=barcode,
                item_name=item_name,
                current_price=current_price,
                uploaded_price=uploaded_price,
                status=row_status,
                apply_selected=apply_selected,
                error_message=error_message,
            )

            db.add(import_row)

        batch.total_rows = len(extracted_rows)
        batch.changed_rows = changed_rows
        batch.unchanged_rows = unchanged_rows
        batch.invalid_rows = invalid_rows

        db.commit()

        return get_price_import_preview(
            db=db,
            batch_id=batch.id,
        )

    except Exception:
        db.rollback()
        raise

    finally:
        await excel_file.close()

def apply_master_price_import(
    db: Session,
    batch_id: int,
) -> PriceImportBatch:
    try:
        batch = db.scalar(
            select(PriceImportBatch)
            .where(PriceImportBatch.id == batch_id)
            .with_for_update()
        )

        if not batch:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Price import preview not found.",
            )

        if batch.import_scope != "master":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="This import is not a master-price import.",
            )

        if batch.status != "preview":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=(
                    "This import has already been processed. "
                    f"Current status: {batch.status}."
                ),
            )

        import_rows = list(
            db.scalars(
                select(PriceImportRow)
                .where(
                    PriceImportRow.batch_id == batch_id,
                    PriceImportRow.status == "changed",
                    PriceImportRow.apply_selected.is_(True),
                )
                .order_by(
                    PriceImportRow.excel_row_number
                )
                .with_for_update()
            ).all()
        )

        if not import_rows:
            batch.status = "applied"
            batch.applied_at = datetime.now(timezone.utc)

            db.commit()

            return get_price_import_preview(
                db=db,
                batch_id=batch.id,
            )

        product_ids = {
            import_row.product_id
            for import_row in import_rows
            if import_row.product_id is not None
        }

        products = list(
            db.scalars(
                select(Product)
                .where(Product.id.in_(product_ids))
                .with_for_update()
            ).all()
        )

        products_by_id = {
            product.id: product
            for product in products
        }

        conflicts: list[dict[str, Any]] = []

        for import_row in import_rows:
            product = products_by_id.get(
                import_row.product_id
            )

            if product is None:
                conflicts.append(
                    {
                        "excel_row_number": (
                            import_row.excel_row_number
                        ),
                        "barcode": import_row.barcode,
                        "reason": "Product no longer exists.",
                    }
                )
                continue

            if product.barcode != import_row.barcode:
                conflicts.append(
                    {
                        "excel_row_number": (
                            import_row.excel_row_number
                        ),
                        "barcode": import_row.barcode,
                        "reason": (
                            "Product barcode changed after preview."
                        ),
                    }
                )
                continue

            if (
                import_row.current_price is None
                or import_row.uploaded_price is None
            ):
                conflicts.append(
                    {
                        "excel_row_number": (
                            import_row.excel_row_number
                        ),
                        "barcode": import_row.barcode,
                        "reason": (
                            "Preview price information is incomplete."
                        ),
                    }
                )
                continue

            database_price = Decimal(
                product.master_price
            ).quantize(TWO_DECIMAL_PLACES)

            preview_price = Decimal(
                import_row.current_price
            ).quantize(TWO_DECIMAL_PLACES)

            if database_price != preview_price:
                conflicts.append(
                    {
                        "excel_row_number": (
                            import_row.excel_row_number
                        ),
                        "barcode": import_row.barcode,
                        "preview_price": str(preview_price),
                        "current_database_price": str(
                            database_price
                        ),
                        "reason": (
                            "Master price changed after preview."
                        ),
                    }
                )

        if conflicts:
            db.rollback()

            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail={
                    "message": (
                        "Prices were not applied because some "
                        "products changed after the preview. "
                        "Upload the Excel file again to create "
                        "a fresh preview."
                    ),
                    "total_conflicts": len(conflicts),
                    "conflicts": conflicts[:50],
                },
            )

        for import_row in import_rows:
            product = products_by_id[
                import_row.product_id
            ]

            product.master_price = (
                import_row.uploaded_price
            )

            import_row.status = "applied"

        batch.status = "applied"
        batch.applied_at = datetime.now(timezone.utc)

        # Existing branch overrides remain untouched.
        db.commit()

        return get_price_import_preview(
            db=db,
            batch_id=batch.id,
        )

    except HTTPException:
        db.rollback()
        raise

    except Exception:
        db.rollback()
        raise


async def create_branch_price_preview(
    db: Session,
    branch_id: int,
    excel_file: UploadFile,
) -> PriceImportBatch:
    try:
        get_branch_by_id(
            db=db,
            branch_id=branch_id,
        )

        original_filename = Path(
            excel_file.filename or "branch-price-import.xlsx"
        ).name

        file_extension = Path(
            original_filename
        ).suffix.lower()

        if file_extension not in {".xlsx", ".xls"}:
            raise HTTPException(
                status_code=(
                    status.HTTP_415_UNSUPPORTED_MEDIA_TYPE
                ),
                detail=(
                    "Only .xlsx and .xls Excel files are supported."
                ),
            )

        file_content = await excel_file.read(
            MAXIMUM_FILE_SIZE + 1
        )

        if not file_content:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="The uploaded Excel file is empty.",
            )

        if len(file_content) > MAXIMUM_FILE_SIZE:
            raise HTTPException(
                status_code=(
                    status.HTTP_413_REQUEST_ENTITY_TOO_LARGE
                ),
                detail="Excel file size cannot exceed 10 MB.",
            )

        extracted_rows = extract_excel_rows(
            file_content=file_content,
            file_extension=file_extension,
        )

        if not extracted_rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No product rows were found in the Excel file.",
            )

        barcode_counts = Counter(
            row["barcode"]
            for row in extracted_rows
            if row["barcode"]
        )

        valid_barcodes = {
            row["barcode"]
            for row in extracted_rows
            if (
                row["barcode"]
                and row["item_name"]
                and row["uploaded_price"] is not None
                and barcode_counts[row["barcode"]] == 1
            )
        }

        products_by_barcode = get_products_by_barcodes(
            db=db,
            barcodes=valid_barcodes,
        )

        product_ids = {
            product.id
            for product in products_by_barcode.values()
        }

        overrides_by_product_id = {}

        if product_ids:
            existing_overrides = db.scalars(
                select(BranchPriceOverride).where(
                    BranchPriceOverride.branch_id == branch_id,
                    BranchPriceOverride.product_id.in_(
                        product_ids
                    ),
                )
            ).all()

            overrides_by_product_id = {
                price_override.product_id: price_override
                for price_override in existing_overrides
            }

        batch = PriceImportBatch(
            import_scope="branch",
            branch_id=branch_id,
            original_filename=original_filename[:255],
            status="preview",
        )

        db.add(batch)
        db.flush()

        changed_rows = 0
        unchanged_rows = 0
        invalid_rows = 0

        for extracted_row in extracted_rows:
            barcode = extracted_row["barcode"]
            item_name = extracted_row["item_name"]
            uploaded_price = extracted_row["uploaded_price"]

            row_status = "invalid"
            error_message = None
            apply_selected = False
            product = None
            current_price = None

            validation_errors = []

            if barcode is None:
                validation_errors.append(
                    "Invalid or missing barcode."
                )

            if item_name is None:
                validation_errors.append(
                    "Invalid or missing item name."
                )

            if uploaded_price is None:
                validation_errors.append(
                    "Invalid or missing price."
                )

            if (
                barcode
                and barcode_counts[barcode] > 1
            ):
                validation_errors.append(
                    "Duplicate barcode exists in the uploaded file."
                )

            if validation_errors:
                error_message = " ".join(
                    validation_errors
                )

                invalid_rows += 1

            else:
                product = products_by_barcode.get(
                    barcode
                )

                if product is None:
                    row_status = "product_not_found"
                    error_message = (
                        "No existing product was found "
                        "for this barcode."
                    )

                    invalid_rows += 1

                else:
                    existing_override = (
                        overrides_by_product_id.get(
                            product.id
                        )
                    )

                    if existing_override:
                        effective_price = (
                            existing_override.override_price
                        )
                    else:
                        effective_price = (
                            product.master_price
                        )

                    current_price = Decimal(
                        effective_price
                    ).quantize(TWO_DECIMAL_PLACES)

                    if current_price == uploaded_price:
                        row_status = "unchanged"
                        unchanged_rows += 1

                    else:
                        row_status = "changed"
                        apply_selected = True
                        changed_rows += 1

            import_row = PriceImportRow(
                batch_id=batch.id,
                excel_row_number=(
                    extracted_row["excel_row_number"]
                ),
                product_id=(
                    product.id if product else None
                ),
                barcode=barcode,
                item_name=item_name,
                current_price=current_price,
                uploaded_price=uploaded_price,
                status=row_status,
                apply_selected=apply_selected,
                error_message=error_message,
            )

            db.add(import_row)

        batch.total_rows = len(extracted_rows)
        batch.changed_rows = changed_rows
        batch.unchanged_rows = unchanged_rows
        batch.invalid_rows = invalid_rows

        db.commit()

        return get_price_import_preview(
            db=db,
            batch_id=batch.id,
        )

    except Exception:
        db.rollback()
        raise

    finally:
        await excel_file.close()


def apply_branch_price_import(
    db: Session,
    batch_id: int,
) -> PriceImportBatch:
    try:
        batch = db.scalar(
            select(PriceImportBatch)
            .where(PriceImportBatch.id == batch_id)
            .with_for_update()
        )

        if not batch:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Price import preview not found.",
            )

        if batch.import_scope != "branch":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="This import is not a branch-price import.",
            )

        if batch.branch_id is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Branch information is missing from this import.",
            )

        if batch.status != "preview":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=(
                    "This import has already been processed. "
                    f"Current status: {batch.status}."
                ),
            )

        branch = get_branch_by_id(
            db=db,
            branch_id=batch.branch_id,
        )

        import_rows = list(
            db.scalars(
                select(PriceImportRow)
                .where(
                    PriceImportRow.batch_id == batch_id,
                    PriceImportRow.status == "changed",
                    PriceImportRow.apply_selected.is_(True),
                )
                .order_by(
                    PriceImportRow.excel_row_number
                )
                .with_for_update()
            ).all()
        )

        if not import_rows:
            batch.status = "applied"
            batch.applied_at = datetime.now(timezone.utc)

            db.commit()

            return get_price_import_preview(
                db=db,
                batch_id=batch.id,
            )

        product_ids = {
            import_row.product_id
            for import_row in import_rows
            if import_row.product_id is not None
        }

        products = list(
            db.scalars(
                select(Product)
                .where(Product.id.in_(product_ids))
                .with_for_update()
            ).all()
        )

        products_by_id = {
            product.id: product
            for product in products
        }

        existing_overrides = list(
            db.scalars(
                select(BranchPriceOverride)
                .where(
                    BranchPriceOverride.branch_id
                    == branch.id,
                    BranchPriceOverride.product_id.in_(
                        product_ids
                    ),
                )
                .with_for_update()
            ).all()
        )

        overrides_by_product_id = {
            price_override.product_id: price_override
            for price_override in existing_overrides
        }

        conflicts: list[dict[str, Any]] = []

        for import_row in import_rows:
            product = products_by_id.get(
                import_row.product_id
            )

            if product is None:
                conflicts.append(
                    {
                        "excel_row_number": (
                            import_row.excel_row_number
                        ),
                        "barcode": import_row.barcode,
                        "reason": "Product no longer exists.",
                    }
                )
                continue

            if product.barcode != import_row.barcode:
                conflicts.append(
                    {
                        "excel_row_number": (
                            import_row.excel_row_number
                        ),
                        "barcode": import_row.barcode,
                        "reason": (
                            "Product barcode changed after preview."
                        ),
                    }
                )
                continue

            if (
                import_row.current_price is None
                or import_row.uploaded_price is None
            ):
                conflicts.append(
                    {
                        "excel_row_number": (
                            import_row.excel_row_number
                        ),
                        "barcode": import_row.barcode,
                        "reason": (
                            "Preview price information is incomplete."
                        ),
                    }
                )
                continue

            existing_override = (
                overrides_by_product_id.get(product.id)
            )

            if existing_override:
                current_effective_price = Decimal(
                    existing_override.override_price
                ).quantize(TWO_DECIMAL_PLACES)
            else:
                current_effective_price = Decimal(
                    product.master_price
                ).quantize(TWO_DECIMAL_PLACES)

            preview_price = Decimal(
                import_row.current_price
            ).quantize(TWO_DECIMAL_PLACES)

            if current_effective_price != preview_price:
                conflicts.append(
                    {
                        "excel_row_number": (
                            import_row.excel_row_number
                        ),
                        "barcode": import_row.barcode,
                        "preview_price": str(preview_price),
                        "current_branch_price": str(
                            current_effective_price
                        ),
                        "reason": (
                            "Branch price changed after preview."
                        ),
                    }
                )

        if conflicts:
            db.rollback()

            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail={
                    "message": (
                        "Branch prices were not applied because "
                        "some prices changed after the preview. "
                        "Upload the Excel file again to create "
                        "a fresh preview."
                    ),
                    "total_conflicts": len(conflicts),
                    "conflicts": conflicts[:50],
                },
            )

        for import_row in import_rows:
            product = products_by_id[
                import_row.product_id
            ]

            uploaded_price = Decimal(
                import_row.uploaded_price
            ).quantize(TWO_DECIMAL_PLACES)

            master_price = Decimal(
                product.master_price
            ).quantize(TWO_DECIMAL_PLACES)

            existing_override = (
                overrides_by_product_id.get(product.id)
            )

            if uploaded_price == master_price:
                # Uploaded price matches master price,
                # therefore this branch should follow master.
                if existing_override:
                    db.delete(existing_override)

            elif existing_override:
                existing_override.override_price = (
                    uploaded_price
                )

            else:
                new_override = BranchPriceOverride(
                    branch_id=branch.id,
                    product_id=product.id,
                    override_price=uploaded_price,
                )

                db.add(new_override)

            import_row.status = "applied"

        batch.status = "applied"
        batch.applied_at = datetime.now(timezone.utc)

        db.commit()

        return get_price_import_preview(
            db=db,
            batch_id=batch.id,
        )

    except HTTPException:
        db.rollback()
        raise

    except Exception:
        db.rollback()
        raise