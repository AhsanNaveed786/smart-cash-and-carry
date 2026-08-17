import hashlib
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import or_, select
from sqlalchemy.orm import Session, selectinload

from models import (
    Admin,
    Branch,
    DataExport,
    Order,
    OrderExportItem,
    OrderStatusHistory,
    Product,
    RevenueOrderLedger,
)
from services.admin_pricing_service import build_product_price_row
from services.order_service import get_order_date_boundaries
from services.rbac_service import create_admin_audit_log
from services.revenue_service import record_completed_order_revenue


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="173F35",
)
HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)
MAXIMUM_EXPORT_ROWS = 100_000


def style_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(
            column_cells[0].column
        )
        maximum_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[column_letter].width = min(
            max(maximum_length + 2, 12),
            45,
        )


def workbook_to_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def excel_datetime(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def get_branch_or_404(
    db: Session,
    branch_id: int,
) -> Branch:
    branch = db.get(Branch, branch_id)
    if branch is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Branch not found.",
        )
    return branch


def export_orders_workbook(
    db: Session,
    super_admin: Admin,
    branch_id: int | None = None,
    order_status: str | None = None,
    created_from: date | None = None,
    created_to: date | None = None,
    ip_address: str | None = None,
) -> tuple[bytes, DataExport]:
    conditions = []
    if branch_id is not None:
        get_branch_or_404(db, branch_id)
        conditions.append(Order.branch_id == branch_id)
    if order_status is not None:
        conditions.append(Order.status == order_status)

    start_at, end_before = get_order_date_boundaries(
        created_from=created_from,
        created_to=created_to,
    )
    if start_at is not None:
        conditions.append(Order.created_at >= start_at)
    if end_before is not None:
        conditions.append(Order.created_at < end_before)

    statement = (
        select(Order)
        .options(
            selectinload(Order.items),
            selectinload(Order.branch),
        )
        .order_by(Order.created_at, Order.id)
        .limit(MAXIMUM_EXPORT_ROWS + 1)
    )
    if conditions:
        statement = statement.where(*conditions)

    orders = list(db.scalars(statement).all())
    if len(orders) > MAXIMUM_EXPORT_ROWS:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=(
                "Export contains more than 100,000 orders. "
                "Use a smaller date range."
            ),
        )

    workbook = Workbook()
    orders_sheet = workbook.active
    orders_sheet.title = "Orders"
    orders_sheet.append(
        [
            "Order Number",
            "Branch",
            "Customer Name",
            "Phone Number",
            "Customer Email",
            "Fulfillment",
            "Channel",
            "Payment Method",
            "Status",
            "Subtotal",
            "Delivery Fee",
            "Total Amount",
            "Delivery Address",
            "City",
            "Notes",
            "Process After",
            "Created At",
            "Updated At",
        ]
    )

    items_sheet = workbook.create_sheet("Order Items")
    items_sheet.append(
        [
            "Order Number",
            "Product ID",
            "Variant ID",
            "Product Name",
            "Variant Name",
            "SKU",
            "Quantity",
            "Unit Price",
            "Line Total",
        ]
    )

    total_amount = Decimal("0.00")
    for order in orders:
        total_amount += Decimal(order.total_amount)
        orders_sheet.append(
            [
                order.order_number,
                order.branch.name,
                order.customer_name,
                order.phone_number,
                order.customer_email,
                order.fulfillment_method,
                order.order_channel,
                order.payment_method,
                order.status,
                float(order.subtotal),
                float(order.delivery_fee),
                float(order.total_amount),
                order.delivery_address,
                order.city,
                order.notes,
                excel_datetime(order.process_after),
                excel_datetime(order.created_at),
                excel_datetime(order.updated_at),
            ]
        )
        for item in order.items:
            items_sheet.append(
                [
                    order.order_number,
                    item.product_id,
                    item.variant_id,
                    item.product_name,
                    item.variant_name,
                    item.sku,
                    item.quantity,
                    float(item.unit_price),
                    float(item.line_total),
                ]
            )

    style_worksheet(orders_sheet)
    style_worksheet(items_sheet)
    content = workbook_to_bytes(workbook)
    timestamp = datetime.now(timezone.utc).strftime(
        "%Y%m%d_%H%M%S"
    )
    file_name = f"smart_cash_orders_{timestamp}.xlsx"

    filters_snapshot = {
        "branch_id": branch_id,
        "order_status": order_status,
        "created_from": (
            created_from.isoformat() if created_from else None
        ),
        "created_to": (
            created_to.isoformat() if created_to else None
        ),
    }

    try:
        export_record = DataExport(
            export_type="orders",
            status="completed",
            branch_id=branch_id,
            created_by_admin_id=super_admin.id,
            file_name=file_name,
            file_sha256=hashlib.sha256(content).hexdigest(),
            filters_snapshot=filters_snapshot,
            record_count=len(orders),
            total_amount=total_amount,
            allows_order_deletion=(
                bool(orders)
                and order_status == "completed"
                and all(order.status == "completed" for order in orders)
            ),
        )
        db.add(export_record)
        db.flush()

        for order in orders:
            db.add(
                OrderExportItem(
                    export_id=export_record.id,
                    order_id=order.id,
                    order_number=order.order_number,
                    branch_id=order.branch_id,
                    status_at_export=order.status,
                    total_amount=order.total_amount,
                    order_created_at=order.created_at,
                )
            )

        create_admin_audit_log(
            db=db,
            action="orders.exported",
            actor_admin_id=super_admin.id,
            details={
                "export_id": export_record.id,
                "record_count": len(orders),
                "filters": filters_snapshot,
                "allows_order_deletion": (
                    export_record.allows_order_deletion
                ),
            },
            ip_address=ip_address,
        )
        db.commit()
        db.refresh(export_record)
        return content, export_record
    except Exception:
        db.rollback()
        raise


def export_products_workbook(
    db: Session,
    super_admin: Admin,
    branch_id: int | None = None,
    search: str | None = None,
    active_only: bool = False,
    ip_address: str | None = None,
) -> tuple[bytes, DataExport]:
    if branch_id is not None:
        branches = [get_branch_or_404(db, branch_id)]
    else:
        branches = list(
            db.scalars(
                select(Branch)
                .where(Branch.is_active.is_(True))
                .order_by(Branch.id)
            ).all()
        )

    conditions = []
    if search:
        pattern = f"%{search.strip()}%"
        conditions.append(
            or_(
                Product.name.ilike(pattern),
                Product.barcode.ilike(pattern),
            )
        )
    if active_only:
        conditions.append(Product.is_active.is_(True))

    statement = (
        select(Product)
        .options(
            selectinload(Product.category),
            selectinload(Product.price_overrides),
        )
        .order_by(Product.name, Product.id)
        .limit(MAXIMUM_EXPORT_ROWS + 1)
    )
    if conditions:
        statement = statement.where(*conditions)
    products = list(db.scalars(statement).all())
    if len(products) > MAXIMUM_EXPORT_ROWS:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Product export is too large.",
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Product Prices"

    headers = [
        "Product ID",
        "Barcode",
        "Product Name",
        "Category",
        "Master Price",
        "Active",
        "Same Price On All Selected Branches",
        "Different Price Branches",
    ]
    for branch in branches:
        headers.extend(
            [
                f"{branch.name} Price",
                f"{branch.name} Source",
            ]
        )
    worksheet.append(headers)

    for product in products:
        price_row = build_product_price_row(
            product=product,
            branches=branches,
        )
        row = [
            product.id,
            product.barcode,
            product.name,
            product.category.name,
            float(product.master_price),
            "Yes" if product.is_active else "No",
            (
                "Yes"
                if price_row["same_price_on_all_branches"]
                else "No"
            ),
            ", ".join(price_row["different_branch_names"]),
        ]
        for branch_price in price_row["branch_prices"]:
            row.extend(
                [
                    float(branch_price["effective_price"]),
                    branch_price["price_source"],
                ]
            )
        worksheet.append(row)

    style_worksheet(worksheet)
    content = workbook_to_bytes(workbook)
    timestamp = datetime.now(timezone.utc).strftime(
        "%Y%m%d_%H%M%S"
    )
    file_name = f"smart_cash_product_prices_{timestamp}.xlsx"
    filters_snapshot = {
        "branch_id": branch_id,
        "search": search,
        "active_only": active_only,
    }

    try:
        export_record = DataExport(
            export_type="products",
            status="completed",
            branch_id=branch_id,
            created_by_admin_id=super_admin.id,
            file_name=file_name,
            file_sha256=hashlib.sha256(content).hexdigest(),
            filters_snapshot=filters_snapshot,
            record_count=len(products),
            total_amount=Decimal("0.00"),
            allows_order_deletion=False,
        )
        db.add(export_record)
        db.flush()
        create_admin_audit_log(
            db=db,
            action="products.exported",
            actor_admin_id=super_admin.id,
            details={
                "export_id": export_record.id,
                "record_count": len(products),
                "filters": filters_snapshot,
            },
            ip_address=ip_address,
        )
        db.commit()
        db.refresh(export_record)
        return content, export_record
    except Exception:
        db.rollback()
        raise


def list_export_records(
    db: Session,
    export_type: str | None = None,
    skip: int = 0,
    limit: int = 100,
) -> list[DataExport]:
    statement = select(DataExport).order_by(
        DataExport.created_at.desc(),
        DataExport.id.desc(),
    )
    if export_type is not None:
        statement = statement.where(
            DataExport.export_type == export_type
        )
    return list(
        db.scalars(
            statement.offset(skip).limit(limit)
        ).all()
    )


def delete_orders_from_export(
    db: Session,
    export_id: int,
    super_admin: Admin,
    ip_address: str | None = None,
) -> dict:
    export_record = db.scalar(
        select(DataExport)
        .options(selectinload(DataExport.order_items))
        .where(DataExport.id == export_id)
        .with_for_update()
    )
    if export_record is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Export record not found.",
        )
    if (
        export_record.export_type != "orders"
        or export_record.status != "completed"
        or not export_record.allows_order_deletion
    ):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "Only a successful completed-orders export can be "
                "used for permanent deletion."
            ),
        )

    order_ids = [
        item.order_id
        for item in export_record.order_items
        if item.order_id is not None
        and item.deleted_at is None
        and item.status_at_export == "completed"
    ]
    if not order_ids:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="The exported orders were already deleted.",
        )

    orders = list(
        db.scalars(
            select(Order)
            .where(Order.id.in_(order_ids))
            .with_for_update()
        ).all()
    )
    if len(orders) != len(set(order_ids)):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Some exported orders no longer exist.",
        )
    if any(order.status != "completed" for order in orders):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Every order must still have completed status.",
        )

    deletion_time = datetime.now(timezone.utc)
    order_numbers = [order.order_number for order in orders]
    status_history_count = len(
        db.scalars(
            select(OrderStatusHistory.id).where(
                OrderStatusHistory.order_number.in_(order_numbers)
            )
        ).all()
    )

    try:
        for order in orders:
            revenue_exists = db.scalar(
                select(RevenueOrderLedger.id).where(
                    RevenueOrderLedger.order_number
                    == order.order_number
                )
            )
            if revenue_exists is None:
                completed_at = db.scalar(
                    select(OrderStatusHistory.created_at)
                    .where(
                        OrderStatusHistory.order_number
                        == order.order_number,
                        OrderStatusHistory.new_status
                        == "completed",
                    )
                    .order_by(OrderStatusHistory.created_at.desc())
                    .limit(1)
                )
                record_completed_order_revenue(
                    db=db,
                    order=order,
                    completed_at=(
                        completed_at
                        or order.updated_at
                        or deletion_time
                    ),
                )

        related_export_items = db.scalars(
            select(OrderExportItem).where(
                OrderExportItem.order_number.in_(order_numbers),
                OrderExportItem.deleted_at.is_(None),
            )
        ).all()
        for export_item in related_export_items:
            export_item.deleted_at = deletion_time
            export_item.deleted_by_admin_id = super_admin.id

        for order in orders:
            db.delete(order)

        deleted_count = len(orders)
        export_record.deleted_order_count += deleted_count
        export_record.orders_deleted_at = deletion_time

        create_admin_audit_log(
            db=db,
            action="orders.deleted_after_export",
            actor_admin_id=super_admin.id,
            details={
                "export_id": export_record.id,
                "deleted_orders": deleted_count,
                "order_numbers": order_numbers,
            },
            ip_address=ip_address,
        )
        db.commit()
        return {
            "message": "Exported completed orders permanently deleted.",
            "export_id": export_record.id,
            "deleted_orders": deleted_count,
            "revenue_records_preserved": deleted_count,
            "status_history_preserved": status_history_count,
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise
